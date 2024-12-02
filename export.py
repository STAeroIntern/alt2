from library import *
import plot
from openpyxl.styles import Alignment

def run(Data_Lib):
    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Figures"

    # Set up headers
    ws.append(['Date', 'Time', 'Description', 'UAV ID', 'Remark', "Image"])

    # In-memory list to store images for Excel
    img_buffers = []

    for files in Data_Lib:
        try:
            for Option in ["UAV vs Pri Baro", "UAV vs Sec Baro", "Pri Baro vs Sec Baro"]:
                # Generate the plot for each file and option and return the figure
                fig, status = plot.run(Data_Lib[files], Option, toggle=0)

                # Create an in-memory buffer for the image based on status
                img_buffer = io.BytesIO()
                if status == 10: 
                    fig.write_image(img_buffer, format='jpg')
                    img_name = f"{files}_{Option}_{str(Data_Lib[files]['UAV ID'].iloc[0])}_Below Cross-Check Offset.jpg"
                elif status == 11:
                    fig.write_image(img_buffer, format='jpg')
                    img_name = f"{files}_{Option}_{str(Data_Lib[files]['UAV ID'].iloc[0])}_Above Cross-Check Offset.jpg"
                elif status == 111:
                    fig.write_image(img_buffer, format='jpg')
                    img_name = f"{files}_{Option}_{str(Data_Lib[files]['UAV ID'].iloc[0])}_Cross-Check Fail.jpg"
                else:
                    fig.write_image(img_buffer, format='jpg')
                    img_name = f"{files}_{Option}_{str(Data_Lib[files]['UAV ID'].iloc[0])}_Below Offset.jpg"
                
                # Add the image buffer to the list (for embedding in Excel later)
                img_buffers.append((img_name, img_buffer))

        except Exception as e:
            st.write(f"Error processing file {files} with option {Option}: {e}")
            continue

    # Create Excel in-memory
    img_column_width = 100
    for row_index, (img_name, img_buffer) in enumerate(img_buffers, start=2):
        try:
            # Extract the file name and split
            filename_without_extension = os.path.splitext(os.path.basename(img_name))[0]
            filename_parts = filename_without_extension.split('_')

            # Add filename parts into separate columns
            for col_index, part in enumerate(filename_parts, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=part)
                cell.alignment = Alignment(horizontal='left', vertical='top')  # Top-left alignment

            # Insert the image into the Excel sheet from the in-memory buffer
            img_buffer.seek(0)  # Ensure buffer is at the start
            excel_img = Image(img_buffer)
            cell_address = f"F{row_index}"
            ws.add_image(excel_img, cell_address)

            # Set row height to accommodate the image
            ws.row_dimensions[row_index].height = excel_img.height

        except Exception as e:
            st.write(f"Error inserting image {img_name} into Excel: {e}")
            continue

    # Set column dimensions for better formatting
    ws.column_dimensions["A"].width = 15  # Date
    ws.column_dimensions["B"].width = 15  # Time
    ws.column_dimensions["C"].width = 20  # Description
    ws.column_dimensions["D"].width = 15  # UAV ID
    ws.column_dimensions["E"].width = 30  # Remark
    ws.column_dimensions["F"].width = img_column_width  # Image column

    # Save the workbook to an in-memory buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)  # Rewind buffer to the beginning

    # Return the in-memory Excel buffer for download
    return excel_buffer

# def run(Data_Lib):
#     # Create a new workbook and add a worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Figures"

#     # Set up headers
#     ws.append(['Date','Time','Description','UAV ID','Remark',"Image"])

#     for files in Data_Lib:
#         try:
#             for Option in ["UAV vs Pri Baro","UAV vs Sec Baro","Pri Baro vs Sec Baro"]:
#                 #Grab the plot for each files and option and return the figure instead of plotting
#                 fig,status = plot.run(Data_Lib[files],Option,toggle=0)
#                 if status == 10: 
#                     #Write the image file name
#                     fig.write_image(r"D:\alt\save\\"+ str(files) + '_' + Option + '_' + str(Data_Lib[files]['UAV ID'].iloc[0]) + '_' + 'Below Cross-Check Offset' +'.jpg')
#                 elif status == 11:
#                     fig.write_image(r"D:\alt\save\\"+ str(files) + '_' + Option + '_' + str(Data_Lib[files]['UAV ID'].iloc[0]) + '_' + 'Above Cross-Check Offset' +'.jpg')
#                 elif status == 111:
#                     fig.write_image(r"D:\alt\save\\"+ str(files) + '_' + Option + '_' + str(Data_Lib[files]['UAV ID'].iloc[0]) + '_' + 'Cross-Check Fail' +'.jpg')
#                 else:
#                     fig.write_image(r"D:\alt\save\\"+ str(files) + '_' + Option + '_' + str(Data_Lib[files]['UAV ID'].iloc[0]) + '_' + 'Below Offset' +'.jpg')
                
#         except:
#             continue
    
#     # Specify the directory containing the images
#     image_directory = r"D:\alt\save\\"
    
#     # Iterate through the image directory
#     for row_index, img in enumerate(os.listdir(image_directory),start = 2):
#         if img.endswith('.jpg'):
        
#             # Extract filename without extension
#             filename_without_extension = os.path.splitext(os.path.basename(img))[0]
            
#             # Split the filename by '_'
#             filename_parts = filename_without_extension.split('_')
        
#             # Add filename parts into separate columns in the current row
#             for col_index, part in enumerate(filename_parts, start=1):
#                 cell = ws.cell(row=row_index, column=col_index, value=part)
#                 cell.alignment = Alignment(horizontal='left', vertical='top')  # Top-left alignment


#             # Create an image object and specify the full path
#             img_path = os.path.join(image_directory, img)
#             excel_img = Image(img_path)

#             # Insert image in the adjacent cell in column D
#             cell_address = f"F{row_index}"  # Adjust for header row
#             ws.add_image(excel_img, cell_address)

#             # Set row height to accommodate the image
#             ws.row_dimensions[row_index].height = excel_img.height


#     ws.column_dimensions["A"].width = 15  # Filename
#     ws.column_dimensions["B"].width = 15  # Prefix
#     ws.column_dimensions["C"].width = 30  # Suffix
#     ws.column_dimensions["D"].width = 15  # Suffix
#     ws.column_dimensions["E"].width = 15 
#     ws.column_dimensions["F"].width = 100  # Image column

#     # Save workbook to Excel file
#     output_file = r"D:\alt\save\report"+".xlsx"
#     wb.save(output_file)

#     print("Excel file saved successfully!")

